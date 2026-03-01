import openpyxl
import os
import sys
from typing import List, Tuple, Dict, Any
from collections import deque

def getFilepathTargetFile(filePathRootFolder: str, filename: str, fileNumber: str, extension: str) -> str:
    """Constructs a cross-platform path for a specific data file."""
    return os.path.join(filePathRootFolder, f"{filename}{fileNumber}.{extension.upper()}")

def getFilepathTargetFile_MHC(filePathRootFolder: str, filename: str, extension: str = "mhc") -> str:
    """Constructs a cross-platform path for the MHC data file."""
    return os.path.join(filePathRootFolder, f"{filename}.{extension.upper()}")

def getNumBranches(filepathRootFolder: str, MHCfilename: str) -> int:
    """Reads the BRN file to determine the total number of branches in the network."""
    BranchFilePath = getFilepathTargetFile(filepathRootFolder, MHCfilename, "", "brn")
    try:
        with open(BranchFilePath) as file:
            lines = file.readlines()
            if len(lines) > 1:
                return int(lines[1].strip())
    except FileNotFoundError:
        print(f"Error: Branch file not found at {BranchFilePath}")
    except ValueError:
        print(f"Error: Could not parse number of branches in {BranchFilePath}")
    return 0

def load_mhc_data(filepathRootFolder: str, MHCfilename: str) -> Dict[str, float]:
    """Loads MHC node ground levels into memory with robust comma/space parsing."""
    mhc_cache = {}
    filepath = getFilepathTargetFile_MHC(filepathRootFolder, MHCfilename)
    try:
        with open(filepath) as file:
            lines = file.readlines()
            for line in lines:
                line = line.strip()
                if not line or not line.startswith('"'): continue
                row = [r.strip(' "') for r in line.replace(",", " ").split() if r.strip()]
                if len(row) >= 4:
                    node_id = row[0]
                    try:
                        ngl = float(row[3])
                        mhc_cache[node_id.lower()] = ngl
                    except (ValueError, IndexError): continue
    except FileNotFoundError:
        print(f"Error opening MHC file at: {filepath}")
    return mhc_cache

def load_all_inv_files(filepathRootFolder: str, MHCfilename: str) -> Dict[str, List[Dict[str, Any]]]:
    """Loads all INV files into memory with robust delimiter handling."""
    inv_cache = {}
    totalBranches = getNumBranches(filepathRootFolder, MHCfilename)
    for branch_idx in range(1, totalBranches + 1):
        fNumStr = f"{branch_idx:03d}"
        filepath = getFilepathTargetFile(filepathRootFolder, MHCfilename, fNumStr, "inv")
        try:
            with open(filepath) as f:
                lines = f.readlines()[3:]
                data = []
                for line in lines:
                    line = line.strip()
                    if not line: continue
                    row = [r.strip(' "') for r in line.replace(",", " ").split() if r.strip()]
                    if len(row) >= 7:
                        try:
                            data.append({
                                "ch": float(row[0]),
                                "il": float(row[1]),
                                "node": row[2],
                                "dia": float(row[3]),
                                "pipe_type": row[6]
                            })
                        except (ValueError, IndexError): continue
                inv_cache[fNumStr] = data
        except FileNotFoundError: continue
    return inv_cache

def load_all_ngl_files(filepathRootFolder: str, MHCfilename: str) -> Dict[str, List[Dict[str, float]]]:
    """Loads all NGL files into memory, supporting both .NGL and .ngl extensions."""
    ngl_cache = {}
    totalBranches = getNumBranches(filepathRootFolder, MHCfilename)
    for branch_idx in range(1, totalBranches + 1):
        fNumStr = f"{branch_idx:03d}"
        found = False
        for ext in ["ngl", "NGL"]:
            filepath = getFilepathTargetFile(filepathRootFolder, MHCfilename, fNumStr, ext)
            if os.path.exists(filepath):
                try:
                    with open(filepath) as f:
                        lines = f.readlines()[4:]
                        data = []
                        for line in lines:
                            line = line.strip()
                            if not line: continue
                            row = [r.strip() for r in line.replace(",", " ").split() if r.strip()]
                            if len(row) >= 2:
                                try:
                                    data.append({"ch": float(row[0]), "ngl": float(row[1])})
                                except: continue
                        if data:
                            ngl_cache[fNumStr] = data
                            found = True
                            break
                except: continue
    return ngl_cache

def get_ngl_at_chainage(chainage: float, ngl_pts: List[Dict[str, float]]) -> float:
    """Linearly interpolates NGL level at a given chainage."""
    if not ngl_pts: return 0.0
    pts = sorted(ngl_pts, key=lambda x: x["ch"])
    if chainage <= pts[0]["ch"]: return pts[0]["ngl"]
    if chainage >= pts[-1]["ch"]: return pts[-1]["ngl"]
    for i in range(len(pts) - 1):
        p1, p2 = pts[i], pts[i+1]
        if p1["ch"] <= chainage <= p2["ch"]:
            dx = p2["ch"] - p1["ch"]
            if dx < 0.0001: return p1["ngl"]
            return p1["ngl"] + (chainage - p1["ch"]) * (p2["ngl"] - p1["ngl"]) / dx
    return pts[-1]["ngl"]

def build_network_graph(inv_cache: Dict[str, List[Dict[str, Any]]]) -> Dict[str, List[Dict[str, Any]]]:
    graph = {}
    for fNumStr, rows in inv_cache.items():
        for i in range(len(rows) - 1):
            u, v = rows[i]["node"].lower(), rows[i+1]["node"].lower()
            if u not in graph: graph[u] = []
            if v not in graph: graph[v] = []
            graph[u].append({"neighbor": v, "branch": fNumStr})
            graph[v].append({"neighbor": u, "branch": fNumStr})
    return graph

def find_path_in_graph(start_node: str, end_node: str, graph: Dict[str, List[Dict[str, Any]]]) -> List[Tuple[str, str, str]]:
    s_node_l, e_node_l = start_node.lower(), end_node.lower()
    if s_node_l not in graph or e_node_l not in graph: return []
    queue = deque([(s_node_l, [])])
    visited = {s_node_l}
    while queue:
        curr, path = queue.popleft()
        if curr == e_node_l:
            if not path: return []
            consolidated = []
            c_branch, c_start, c_end = path[0]["branch"], path[0]["from"], path[0]["to"]
            for i in range(1, len(path)):
                if path[i]["branch"] == c_branch: c_end = path[i]["to"]
                else:
                    consolidated.append((c_branch, c_start, c_end))
                    c_branch, c_start, c_end = path[i]["branch"], path[i]["from"], path[i]["to"]
            consolidated.append((c_branch, c_start, c_end))
            return consolidated
        for edge in graph.get(curr, []):
            if edge["neighbor"] not in visited:
                visited.add(edge["neighbor"])
                queue.append((edge["neighbor"], path + [{"from": curr, "to": edge["neighbor"], "branch": edge["branch"]}]))
    return []

def get_branches(root_folder: str, mhc_filename: str, mode: str, custom_branches: List[str] = None) -> List[List[Tuple[str, str, str]]]:
    inv_cache = load_all_inv_files(root_folder, mhc_filename)
    graph = build_network_graph(inv_cache)
    paths = []
    if mode == "modelled":
        for i in range(1, getNumBranches(root_folder, mhc_filename) + 1):
            paths.append([(f"{i:03d}", None, None)])
    elif mode == "customize" and custom_branches:
        raw = []
        for item in custom_branches: raw.extend([x.strip() for x in item.split(",") if x.strip()])
        for pair in raw:
            if " - " in pair:
                s, e = [x.strip() for x in pair.split(" - ")]
                path = find_path_in_graph(s, e, graph)
                if path: paths.append(path)
                else: raise ValueError(f"Path not found: {s} to {e}")
            else: paths.append([(pair.zfill(3), None, None)])
    return paths

def transferData(paths_list, MHCfilename, filepathRootFolder):
    inv_cache = load_all_inv_files(filepathRootFolder, MHCfilename)
    ngl_cache = load_all_ngl_files(filepathRootFolder, MHCfilename)
    mhc_cache = load_mhc_data(filepathRootFolder, MHCfilename)
    all_labels, all_every_ch, all_every_ngl, all_every_il, all_diameters, all_pipe_types, all_slopes = [], [], [], [], [], [], []
    
    for segments in paths_list:
        m_lbl, m_ch, m_ngl, m_il, m_dia, m_pt, m_slopes = [], [], [], [], [], [], []
        for fNumStr, s_node, e_node in segments:
            rows = inv_cache.get(fNumStr, [])
            ngl_pts = ngl_cache.get(fNumStr, [])
            if not rows or not ngl_pts: continue
            pts = []
            for r in rows:
                if pts and abs(r["ch"] - pts[-1]["ch"]) < 0.0001:
                    if -500 < r["il"] < 500: pts[-1]["il_out"] += r["il"]
                    else: pts[-1]["il_out"] = r["il"]
                    continue
                pts.append({"ch": r["ch"], "il_in": r["il"], "il_out": r["il"], "node": r["node"], "dia": r["dia"], "pt": r["pipe_type"]})
            
            if s_node and e_node:
                s_l, e_l = s_node.lower(), e_node.lower()
                idx_s = next((i for i, p in enumerate(pts) if p["node"].lower() == s_l), -1)
                idx_e = next((i for i, p in enumerate(pts) if p["node"].lower() == e_l), -1)
                if idx_s == -1 or idx_e == -1: continue
                rev = idx_s > idx_e
                pts = pts[min(idx_s, idx_e):max(idx_s, idx_e)+1]
                if rev: 
                    pts = pts[::-1]
                    for p in pts: p["il_in"], p["il_out"] = p["il_out"], p["il_in"]

            for i in range(len(pts) - 1):
                p1, p2 = pts[i], pts[i+1]
                dist = abs(p2["ch"] - p1["ch"])
                slope = (p1["il_out"] - p2["il_in"]) / dist if dist > 0.0001 else 0.0
                m_dia.append(p1["dia"]); m_pt.append(p1["pt"]); m_slopes.append(slope)
                if not m_lbl: m_lbl.append(p1["node"])
                m_lbl.append(p2["node"])
                
                s_ch, s_ngl, s_il = [], [], []
                sn_s, sn_e = p1["ch"], p2["ch"]
                lo, hi = min(sn_s, sn_e), max(sn_s, sn_e)
                seg_ngl = [p for p in ngl_pts if lo - 0.0001 <= p["ch"] <= hi + 0.0001]
                if sn_s > sn_e: seg_ngl = seg_ngl[::-1]
                for p in seg_ngl:
                    s_ch.append(p["ch"]); s_ngl.append(p["ngl"])
                    s_il.append(round(p1["il_out"] - slope * abs(p["ch"] - sn_s), 3))
                if not s_ch or abs(s_ch[0] - sn_s) > 0.001:
                    s_ch.insert(0, sn_s); s_ngl.insert(0, mhc_cache.get(p1["node"].lower(), get_ngl_at_chainage(sn_s, ngl_pts))); s_il.insert(0, p1["il_out"])
                if abs(s_ch[-1] - sn_e) > 0.001:
                    s_ch.append(sn_e); s_ngl.append(mhc_cache.get(p2["node"].lower(), get_ngl_at_chainage(sn_e, ngl_pts))); s_il.append(p2["il_in"])
                s_il[-1] = p2["il_in"]
                m_ch.append(s_ch); m_ngl.append(s_ngl); m_il.append(s_il)

        if m_lbl:
            clean = [m_lbl[0]]
            for l in m_lbl[1:]:
                if l != clean[-1]: clean.append(l)
            all_labels.append(clean); all_every_ch.append(m_ch); all_every_ngl.append(m_ngl); all_every_il.append(m_il)
            all_diameters.append(m_dia); all_pipe_types.append(m_pt); all_slopes.append(m_slopes)
    return all_labels, all_every_ch, all_every_ngl, all_every_il, all_diameters, all_pipe_types, all_slopes

def OutsideDiameter_Sewer(pipeType: list) -> list:
    OD = []
    for branch in pipeType:
        temp = []
        for seg in branch:
            try: temp.append(float(seg.split(" ")[0].split("mm")[0]))
            except: temp.append(0.0)
        OD.append(temp)
    return OD

def generateSpreadsheet(nodeLabels, pipeTypes, innerDia, outerDia, chainages, NGLs, ILs, output_path):
    HEADERS = ["Node ID", "Pipe Type", "Inner Diameter (m)", "Outside Diameter (m)", "Bedding Depth(m)", "Pipe Thickeness (m)", "Working Space (m)", " ", "Chainage (m)", "Distance (m)","NGL (m)", "IL (m)", "Trench Level (m)", "Trench Depth (m)", "Trench Width (m)", "Excavation (m\u00B3)", " ", "0-1m Deep Excavation (m\u00B3)", "1-2m Deep Excavation (m\u00B3)", "2-3m Deep Excavation (m\u00B3)", "3-4m Deep Excavation (m\u00B3)", "4-5m Deep Excavation (m\u00B3)", "5m-6m Deep Excavation (m\u00B3)",">6m Deep Excavation (m\u00B3)", " ", "0-1m Deep Excavation (m\u00B3)", "1-2m Deep Excavation (m\u00B3)", "2-3m Deep Excavation (m\u00B3)", "3-4m Deep Excavation (m\u00B3)", "4-5m Deep Excavation (m\u00B3)", "5-6m Deep Excavation (m\u00B3)", ">6m Deep Excavation (m\u00B3)" ," ", "Bedding Backfill (m\u00B3)", "Selected Backfill (m\u00B3)", "Additional Selected Backfill (m\u00B3)", "Backfill (m\u00B3)"]
    Summary = ["Excavations", "Hard Rock", " ", "Granular Fill (Reuse)", "Selected Fill (Reuse)", "Granular Fill (Import)", "Selected Fill (Import)", " ", "Total Backfilling", "Imported Backfill Material", "Excess Material"]
    Percentages = [0, 0.2, 0, 0.3, 0.3, 0.7, 0.7, 0, 0, 1, 0]
    wb = openpyxl.Workbook(); ws = wb.active
    align = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
    bold = openpyxl.styles.Font(bold=True)
    for idx, h in enumerate(HEADERS, 1):
        c = ws.cell(2, idx, h); c.alignment = align; c.font = bold
        l = openpyxl.utils.get_column_letter(idx)
        ws.column_dimensions[l].width = 25 if "Excavation" in h or "Pipe Type" in h else 10
        if h == " ": ws.column_dimensions[l].width = 2
    row_idx, skip_rows = 3, []
    for b_idx in range(len(chainages)):
        if b_idx > 0: skip_rows.append(row_idx); row_idx += 1
        labels, b_ch, b_ngl, b_il, b_id, b_od, b_pt = nodeLabels[b_idx], chainages[b_idx], NGLs[b_idx], ILs[b_idx], innerDia[b_idx], outerDia[b_idx], pipeTypes[b_idx]
        start_r = row_idx
        for s_idx in range(len(b_ch)):
            s_ch, s_ngl, s_il = b_ch[s_idx], b_ngl[s_idx], b_il[s_idx]
            for p_idx in range(len(s_ch)):
                if s_idx > 0 and p_idx == 0: continue
                r = row_idx
                if s_idx == 0 and p_idx == 0: ws.cell(r, 1, labels[0])
                elif p_idx == len(s_ch)-1: ws.cell(r, 1, labels[s_idx+1])
                ws.cell(r, 2, b_pt[s_idx]); ws.cell(r, 3, b_id[s_idx]/1000); ws.cell(r, 4, b_od[s_idx]/1000)
                ws.cell(r, 5, f"=IF((D{r}/4)>0.2,0.2, IF((D{r}/4)<0.1,0.1, (D{r}/4)))")
                ws.cell(r, 6, f"=(D{r}-C{r})/2"); ws.cell(r, 7, 0.3); ws.cell(r, 9, s_ch[p_idx])
                ws.cell(r, 10, 0 if r == start_r else f"=IFERROR(I{r}-I{r-1}, 0)")
                ws.cell(r, 11, s_ngl[p_idx]); ws.cell(r, 12, s_il[p_idx])
                ws.cell(r, 13, f"=L{r}-F{r}-E{r}"); ws.cell(r, 14, f"=K{r}-M{r}"); ws.cell(r, 15, f"=D{r}+2*G{r}")
                ws.cell(r, 16, f"=IFERROR(J{r}*O{r}*(N{r}+N{r-1})/2, 0)")
                avg = f"(N{r}+N{r-1})/2"
                for i in range(7):
                    cond = f"AND({avg}>={i}, {avg}<{i+1})" if i < 6 else f"{avg}>=6"
                    ws.cell(r, 18+i, f"=IFERROR(IF(AND({cond}, O{r}<1), J{r}*O{r}*{avg}, 0), 0)")
                    ws.cell(r, 26+i, f"=IFERROR(IF(AND({cond}, O{r}>=1, O{r}<2), J{r}*O{r}*{avg}, 0), 0)")
                ws.cell(r, 34, f"=(J{r}*O{r}*(D{r}+E{r}+F{r})-PI()*((D{r}/2)^2)*J{r})")
                ws.cell(r, 35, f"=J{r}*O{r}*0.2"); ws.cell(r, 37, f"=(P{r}-PI()*((D{r}/2)^2)*J{r})-AH{r}-AI{r}")
                row_idx += 1
    end_r, tot_r = row_idx - 1, row_idx + 1
    for col in [10, 16, 18, 19, 20, 21, 22, 23, 24, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 37]:
        l = openpyxl.utils.get_column_letter(col)
        ws.cell(tot_r, col, f"=SUM({l}3:{l}{end_r})")
    sum_r = tot_r + 2
    for i, m in enumerate(Summary): ws.cell(sum_r + i, 18, m)
    forms = [f"=P{tot_r}", f"=S{sum_r}*T{sum_r+1}", " ", f"=AH{tot_r}*T{sum_r+3}", f"=AI{tot_r}*T{sum_r+4}", f"=AH{tot_r}*T{sum_r+5}", f"=AI{tot_r}*T{sum_r+6}", " ", f"=S{sum_r+3}+S{sum_r+4}", f"=S{sum_r+8}-(S{sum_r}-S{sum_r+1}-S{sum_r+3}-S{sum_r+4})*T{sum_r+9}", f"=S{sum_r+1}"]
    for i, f in enumerate(forms):
        ws.cell(sum_r + i, 19, f)
        if i < len(Percentages) and Percentages[i] > 0: ws.cell(sum_r+i, 20, Percentages[i]).number_format = '0.00%'
    thin = openpyxl.styles.Side(style='thin'); border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(2, tot_r + 1):
        if r in skip_rows: continue
        for c in range(1, ws.max_column + 1):
            if c not in [8, 17, 25, 33]: ws.cell(r, c).border = border
    fill = openpyxl.styles.PatternFill(start_color="FFC7E0BD", end_color="FFC7E0BD", fill_type="solid")
    for r in range(3, end_r + 1):
        if r in skip_rows: continue
        for c in [2, 3, 4, 7, 9, 11, 12]: ws.cell(r, c).fill = fill
    wb.save(output_path)

def run_quantification(network_type, mhc_filename, root_folder, mode, custom_branches=None):
    paths = get_branches(root_folder, mhc_filename, mode, custom_branches)
    if not paths: raise ValueError("No branches selected.")
    lbl, ch, ngl, il, dia, pt, slopes = transferData(paths, mhc_filename, root_folder)
    if not lbl: raise ValueError("No data extracted.")
    out_path = os.path.join(root_folder, f"Quantified_{network_type.lower()}.xlsx")
    generateSpreadsheet(lbl, pt, dia, OutsideDiameter_Sewer(pt), ch, ngl, il, out_path)
    return out_path

if __name__ == "__main__":
    fn = input("Filename: ").lower(); fp = input("Path: ")
    print("1.Finish 2.Modelled 3.Customize")
    mode = "modelled" if input("Enter: ") == "2" else "customize"
    run_quantification("Sewer", fn, fp, mode, [input("Nodes (A - B): ")] if mode=="customize" else None)
